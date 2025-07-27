
# NOTE some utils commands
#    pip freeze > requirements.txt
#    .\.venv\Scripts\Activate.ps1


""" 
# NOTE en caso de que uses un excel para cargar datos con pandas
pip install pandas 
pip install openpyxl 

# NOTE script de ejemplo:
"""

from openpyxl import load_workbook

# Cargar el archivo Excel
wb = load_workbook(filename="datos.xlsx")
sheet = wb.active  # o especifica la hoja: wb["NombreHoja"]

# Encabezados de columnas
columns = ["id", "name", "category"]

# Iterar sobre las filas y crear un diccionario por cada fila
for row in sheet.iter_rows(values_only=True):
    row_dict = {columns[i]: str(row[i]) if row[i] is not None else "" for i in range(len(columns))}
    print(row_dict)


# Resultado
# {'id': '1', 'name': 'computadora', 'category': 'electronica'}
# {'id': '2', 'name': 'celular', 'category': 'telefonia'}


