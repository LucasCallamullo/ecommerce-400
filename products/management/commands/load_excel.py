

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from products import utils
from products.models import Product, PCategory, PSubcategory, PBrand, ProductImage
from users.models import CustomUser

# mys scripts folder for first update
from products.data.load_orders import load_orders_init
from products.data.load_users import load_users_init
from products.data.load_store import load_store_init

# command python manage.py load_data_project
from django.utils.text import slugify
from django.utils.crypto import get_random_string

def unique_slug(base_slug, model):
    slug = base_slug
    while model.objects.filter(slug=slug).exists():
        slug = f"{base_slug}-{get_random_string(4)}"
    return slug


def clean_value(value, zero=False):
    """
        Cleans null or empty values and converts them to None or Zero as appropriate 
        for the database and stores them correctly.

    Args:
        value (any): The value obtained from the data source (e.g., an Excel file).
        zero (bool): Enables returning 0 instead of None for numeric fields.

    Returns:
        Returns a valid value or None or Zero as appropriate
    """
    result = None if not value or value == '' else value

    if zero:
        return 0 if result is None else result
    
    return result
        
def update_main_imagess():
    products = Product.objects.all()

    for product in products:
        images = product.images.all()
        
        if not images:
            print(f'⚠️ Product "{product.name}" has no images.')
            continue

        # Set the first image as main
        first_image = images[0]
        first_image.main_image = True
        first_image.save()

        # Optionally, set all others to False (cleanup)
        for img in images[1:]:
            if img.main_image:
                img.main_image = False
                img.save()

        # Update product's main_image field
        product.main_image = first_image.image_url
        product.save()
        print(f'✔ Main image updated for product "{product.name}"')

from django.contrib.auth.hashers import make_password
from users.models import CustomUser
class Command(BaseCommand):
    help = "To generically load all the necessary data for the models we created "
    help += "in this case, it includes loading Product models from Excel to the database. "
    help += "On the other hand, dictionaries were simply used to create model examples like Store, User, Orders"
        
    def handle(self, *args, **kwargs):
        
        self.stdout.write(self.style.SUCCESS("✔ Desactivado por si sos medio boludo."))
        
        user, created = CustomUser.objects.get_or_create(
            email="admin@gmail.com",
            defaults={
                "password": make_password("1234"),
                "first_name": "Admin",
                "last_name": "SuperAdmin",
                "is_active": True,
                "is_staff": True,
                "is_superuser": True,
                "role": 'admin',
            }
        )
        
        if created:
            print(f'El Super usuario {user.email} Se creo exitosamente')
        
        return

        self.stdout.write(self.style.SUCCESS("✔ Datos iniciales cargados correctamente."))
        # Path to the Excel file
        try:
            file = 'products/data/products_data.xlsx'
            
        except FileNotFoundError:
            print(f'File not found: {file}')
            return None
        
        # Lista de nombres de las columnas en orden
        columns = [
            'id', 'name', 'price', 'available', 'stock', 'category', 'subcategory', 'brand', 
            'discount', 'description', 'image_url', 'image_url2'
        ]

        # Abrimos el archivo Excel
        wb = load_workbook(file)
        ws = wb.active

        # Si querés ignorar el encabezado real del archivo y usar tu lista:
        for fila in ws.iter_rows(min_row=2, values_only=True):  # Salta el encabezado
            fila_dict = dict(zip(columns, fila))  # Empareja nombre-columna con valor
            
            # obtener un name necesario y unico para productos
            name = clean_value(fila_dict.get("name"))
            
            if not name:
                print(f'Product "{fila_dict.get("id")}" does not have a product_name.')
                continue    # no lo va a cargar 
            
            # obtener los valores de cada uno a crear
            category = clean_value(fila_dict.get("category"))
            subcategory = clean_value(fila_dict.get("subcategory"))
            brand = clean_value(fila_dict.get("brand"))
        
            # Create categories, subcategories, brands if they don't exist
            if category is None:
                category_obj = PCategory.get_default_model_or_id(model=True)
            else:
                category_obj, _ = PCategory.objects.get_or_create(name=category)    
            
            if subcategory is None or category is None:
                sub_category_obj = PSubcategory.get_default_model_or_id(model=True)
            else:
                sub_category_obj, _ = PSubcategory.objects.get_or_create(name=subcategory, category=category_obj)    
            
            if brand is None:
                brand_obj = PBrand.get_default_model_or_id(model=True)
            else:
                brand_obj, _ = PBrand.objects.get_or_create(name=brand)    

            # Get the rest of the values from the Excel
            price = clean_value(fila_dict.get("price"), zero=True)
            stock = clean_value(fila_dict.get("stock"), zero=True)
            discount = clean_value(fila_dict.get("discount"), zero=True)
            
            # Return true or false for availability
            available_str = fila_dict.get("available", "").lower()
            available = available_str in ["si", "sí", "yes"]

            description = clean_value(fila_dict.get("description"))
            image_url = clean_value(fila_dict.get("image_url"))
            image_url2 = clean_value(fila_dict.get("image_url2"))
            
            # Normalize the name before creating the product
            normalized_name = utils.normalize_or_None(name)
            slug = unique_slug(slugify(name), Product)  # get the slugified version of the name
            
            # Retrieve or create the product
            product_obj, created = Product.objects.get_or_create(
                name=name,
                slug=slug,
                normalized_name=normalized_name,
                price=price,
                stock=stock,
                discount=discount,
                available=available,
                category=category_obj,
                subcategory=sub_category_obj,
                brand=brand_obj,
                description=description,
            )
            
            # Add image URLs to the products
            cont = 0
            if image_url:
                ProductImage.objects.create(product=product_obj, image_url=image_url)
                cont += 1
        
            if image_url2:
                ProductImage.objects.create(product=product_obj, image_url=image_url2)
                cont += 1
             
            if created:
                print(f'Product {product_obj.name} created successfully with {cont} associated images.')
            else:
                print(f'Product {product_obj.name} updated successfully with {cont} associated images.')
             
        # ================================================================
        # Create other necessary data for the initial project load
        print("=" * 50)
        # This calls the function in scripts/ to load the initial orders
        load_orders_init()
        
        print("=" * 50)
        # This calls the function in scripts/ to load the initial Users
        load_users_init()
        
        print("=" * 50)
        # This calls the function in scripts/ to load the initial Store
        load_store_init()
            
        